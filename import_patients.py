#!/usr/bin/env python3
"""CRM.xlsx → 환자 DB import 스크립트
'신환 내원 경로 개별' 시트의 데이터를 patients 테이블에 삽입
"""
import openpyxl
import json
import subprocess
import uuid
import re
from datetime import datetime

EXCEL_PATH = '/home/user/uploaded_files/CRM.xlsx'
DB_NAME = 'pfm-production'
HOSPITAL_ID = 'h-demo'

# ═══ 매핑 테이블 ═══
# 엑셀 '내원 경로' → DB visit_source 코드
SOURCE_MAP = {
    '환자분소개': 'ref_patient',
    '환자분가족': 'ref_patient',  # 환자 가족도 환자 소개로 분류
    '지인소개': 'ref_acquaintance',
    '직원소개': 'ref_staff',
    '홈페이지': 'online_homepage',
    '홈페이지(DB)': 'online_homepage_db',
    '인스타그램': 'online_insta',
    '당근마켓': 'online_daangn',
    '블로그': 'online_blog',
    '유튜브': 'online_youtube',
    '네이버카페': 'online_cafe',
    '검색': 'online_search',
    '간판, 가까워서': 'walk_sign',
    '간판': 'walk_sign',
    '가까워서': 'walk_near',
    '기타(협력),(광고)': 'online_ad',
    '기타': 'online_etc',
    '두정점환자': 'ref_patient',
    '카카오네비': 'online_search',
}

# 엑셀 '진료 영역' → DB treatment_area 코드
AREA_MAP = {
    '임플란트': 'implant',
    '치아교정': 'orthodontics',
    '심미치료': 'cosmetic',
    '일반진료': 'general',
    '소아치료': 'pediatric',
    '스케일링': 'scaling',
    '틀니': 'denture',
    '기타': 'etc',
}

# 접수자 이름 정규화 (공백 제거)
def normalize_name(name):
    if not name:
        return ''
    return name.strip()

def parse_date(val):
    """날짜 파싱"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s == '#VALUE!':
        return None
    # 다양한 포맷 시도
    for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except:
            pass
    return None

def main():
    print("📂 CRM.xlsx 로딩...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['신환 내원 경로 개별']
    
    records = []
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        year, month, date_val, name, staff, area, source, referrer = row[:8] if len(row) >= 8 else (row + (None,) * (8 - len(row)))
        
        # 이름 없으면 스킵
        patient_name = normalize_name(str(name)) if name else ''
        if not patient_name or patient_name in ('#VALUE!', 'None', ''):
            skipped += 1
            continue
        
        # 날짜 파싱
        visit_date = parse_date(date_val)
        if not visit_date:
            # 연도/월로 대략 추정
            y = str(year).strip() if year else ''
            m = str(month).strip() if month else ''
            if y and y.isdigit() and m and m.isdigit():
                visit_date = f"{y}-{m.zfill(2)}-01"
            else:
                skipped += 1
                continue
        
        # 매핑
        source_str = str(source).strip() if source else ''
        visit_source = SOURCE_MAP.get(source_str, 'online_etc')
        if not source_str:
            visit_source = ''
        
        area_str = str(area).strip() if area else ''
        treatment_area = AREA_MAP.get(area_str, 'etc')
        if not area_str:
            treatment_area = ''
        
        staff_name = normalize_name(str(staff)) if staff else ''
        referrer_name = normalize_name(str(referrer)) if referrer else ''
        
        # visit_source_detail: 원본 경로명 보존
        visit_source_detail = source_str if source_str else ''
        
        records.append({
            'patient_name': patient_name,
            'first_visit_date': visit_date,
            'visit_source': visit_source,
            'visit_source_detail': visit_source_detail,
            'treatment_area': treatment_area,
            'desk_staff': staff_name,
            'assigned_counselor': staff_name,
            'referrer_name': referrer_name,
            'patient_type': 'new',  # 신환 시트이므로
        })
    
    print(f"✅ 파싱 완료: {len(records)}건 (스킵: {skipped}건)")
    
    # 통계 출력
    from collections import Counter
    source_cnt = Counter(r['visit_source'] for r in records if r['visit_source'])
    area_cnt = Counter(r['treatment_area'] for r in records if r['treatment_area'])
    year_cnt = Counter(r['first_visit_date'][:4] for r in records)
    staff_cnt = Counter(r['desk_staff'] for r in records if r['desk_staff'])
    
    print(f"\n📊 내원경로 분포:")
    for k, v in source_cnt.most_common():
        print(f"  {k}: {v}")
    print(f"\n📊 진료영역 분포:")
    for k, v in area_cnt.most_common():
        print(f"  {k}: {v}")
    print(f"\n📊 연도 분포:")
    for k, v in sorted(year_cnt.items()):
        print(f"  {k}: {v}")
    print(f"\n📊 접수자 Top 10:")
    for k, v in staff_cnt.most_common(10):
        print(f"  {k}: {v}")
    
    # ═══ SQL 생성 및 실행 ═══
    print(f"\n🔄 기존 import 데이터 삭제 중...")
    # 기존 import 데이터만 삭제 (created_by='import-crm')
    cmd = f'cd /home/user/webapp && npx wrangler d1 execute {DB_NAME} --local --command="DELETE FROM patients WHERE hospital_id=\'{HOSPITAL_ID}\' AND created_by=\'import-crm\'"'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
    print(f"  삭제 결과: exit={result.returncode}")
    
    # 배치 INSERT (100건씩)
    batch_size = 50
    total_inserted = 0
    
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        sql_parts = []
        
        for r in batch:
            pid = f"pt-{uuid.uuid4().hex[:8]}"
            # SQL 이스케이프
            def esc(s):
                return str(s).replace("'", "''") if s else ''
            
            sql_parts.append(
                f"('{pid}','{HOSPITAL_ID}','','{esc(r['patient_name'])}','','','','{r['patient_type']}',"
                f"'{r['visit_source']}','{esc(r['visit_source_detail'])}','{esc(r['referrer_name'])}',"
                f"'{r['first_visit_date']}','{r['first_visit_date']}',1,"
                f"'{r['treatment_area']}','','{esc(r['assigned_counselor'])}','{esc(r['desk_staff'])}',"
                f"'','','','','','','active','','import-crm')"
            )
        
        sql = (
            "INSERT INTO patients (id, hospital_id, chart_number, patient_name, phone, birth_date, gender, "
            "patient_type, visit_source, visit_source_detail, referrer_name, "
            "first_visit_date, last_visit_date, visit_count, "
            "treatment_area, primary_doctor, assigned_counselor, desk_staff, "
            "visit_reason, address, addr_sido, addr_sigungu, addr_detail, memo, status, kakao_registered, created_by) VALUES "
            + ",".join(sql_parts)
        )
        
        # SQL 파일에 쓰고 실행
        sql_file = f'/tmp/batch_{i}.sql'
        with open(sql_file, 'w') as f:
            f.write(sql + ";\n")
        
        cmd = f'cd /home/user/webapp && npx wrangler d1 execute {DB_NAME} --local --file={sql_file}'
        result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=60)
        
        if result.returncode != 0:
            print(f"  ❌ 배치 {i//batch_size + 1} 실패: {result.stderr[-200:]}")
            # 실패 시 개별 insert 시도
            for j, r in enumerate(batch):
                pid = f"pt-{uuid.uuid4().hex[:8]}"
                def esc(s):
                    return str(s).replace("'", "''") if s else ''
                single_sql = (
                    f"INSERT INTO patients (id, hospital_id, chart_number, patient_name, phone, birth_date, gender, "
                    f"patient_type, visit_source, visit_source_detail, referrer_name, "
                    f"first_visit_date, last_visit_date, visit_count, "
                    f"treatment_area, primary_doctor, assigned_counselor, desk_staff, "
                    f"visit_reason, address, addr_sido, addr_sigungu, addr_detail, memo, status, kakao_registered, created_by) VALUES "
                    f"('{pid}','{HOSPITAL_ID}','','{esc(r['patient_name'])}','','','','{r['patient_type']}',"
                    f"'{r['visit_source']}','{esc(r['visit_source_detail'])}','{esc(r['referrer_name'])}',"
                    f"'{r['first_visit_date']}','{r['first_visit_date']}',1,"
                    f"'{r['treatment_area']}','','{esc(r['assigned_counselor'])}','{esc(r['desk_staff'])}',"
                    f"'','','','','','','active','','import-crm')"
                )
                single_file = f'/tmp/single_{i}_{j}.sql'
                with open(single_file, 'w') as f:
                    f.write(single_sql + ";\n")
                result2 = subprocess.run(
                    f'cd /home/user/webapp && npx wrangler d1 execute {DB_NAME} --local --file={single_file}',
                    shell=True, capture_output=True, text=True, timeout=30
                )
                if result2.returncode == 0:
                    total_inserted += 1
                else:
                    print(f"    개별 실패 [{i+j}]: {r['patient_name']} - {result2.stderr[-100:]}")
        else:
            total_inserted += len(batch)
        
        if (i // batch_size) % 20 == 0:
            print(f"  진행: {min(i + batch_size, len(records))}/{len(records)} ({total_inserted}건 삽입)")
    
    print(f"\n🎉 Import 완료! 총 {total_inserted}건 삽입")
    
    # 검증
    cmd = f'cd /home/user/webapp && npx wrangler d1 execute {DB_NAME} --local --command="SELECT COUNT(*) as c FROM patients WHERE hospital_id=\'{HOSPITAL_ID}\'"'
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=30)
    print(f"📋 DB 총 환자수: {result.stdout}")

if __name__ == '__main__':
    main()
