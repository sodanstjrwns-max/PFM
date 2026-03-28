#!/usr/bin/env python3
"""CRM.xlsx → 환자 DB direct SQLite import (빠른 버전)"""
import openpyxl
import sqlite3
import uuid
from datetime import datetime
from collections import Counter

EXCEL_PATH = '/home/user/uploaded_files/CRM.xlsx'
DB_PATH = '/home/user/webapp/.wrangler/state/v3/d1/miniflare-D1DatabaseObject/d3e041905f05e515c70eeff7a19bb719b7fd5943de6f9e0a4f0dad6a65e8bfec.sqlite'
HOSPITAL_ID = 'h-demo'

# ═══ 매핑 테이블 ═══
SOURCE_MAP = {
    '환자분소개': 'ref_patient',
    '환자분가족': 'ref_patient',
    '지인소개': 'ref_acquaintance',
    '직원소개': 'ref_staff',
    '홈페이지': 'online_homepage',
    '홈페이지(DB)': 'online_homepage_db',
    '인스타그램': 'online_insta',
    '당근마켓': 'online_daangn',
    '블로그': 'online_blog',
    '유튜브': 'online_youtube',
    '네이버카페': 'online_cafe',
    '검색': 'online_search',
    '간판, 가까워서': 'walk_sign',
    '간판': 'walk_sign',
    '가까워서': 'walk_near',
    '기타(협력),(광고)': 'online_ad',
    '기타': 'online_etc',
    '두정점환자': 'ref_patient',
    '카카오네비': 'online_search',
}

AREA_MAP = {
    '임플란트': 'implant',
    '치아교정': 'orthodontics',
    '심미치료': 'cosmetic',
    '일반진료': 'general',
    '소아치료': 'pediatric',
    '스케일링': 'scaling',
    '틀니': 'denture',
    '기타': 'etc',
}

def normalize_name(name):
    if not name:
        return ''
    return str(name).strip()

def parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s == '#VALUE!':
        return None
    for fmt in ['%Y-%m-%d', '%Y.%m.%d', '%Y/%m/%d']:
        try:
            return datetime.strptime(s[:10], fmt).strftime('%Y-%m-%d')
        except:
            pass
    return None

def main():
    print("📂 CRM.xlsx 로딩...")
    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb['신환 내원 경로 개별']
    
    records = []
    skipped = 0
    
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        cols = list(row) + [None] * max(0, 8 - len(row))
        year, month, date_val, name, staff, area, source, referrer = cols[:8]
        
        patient_name = normalize_name(name)
        if not patient_name or patient_name in ('#VALUE!', 'None', ''):
            skipped += 1
            continue
        
        visit_date = parse_date(date_val)
        if not visit_date:
            y = str(year).strip() if year else ''
            m = str(month).strip() if month else ''
            if y and y.replace('.0','').isdigit() and m and m.replace('.0','').isdigit():
                visit_date = f"{int(float(y))}-{str(int(float(m))).zfill(2)}-01"
            else:
                skipped += 1
                continue
        
        source_str = normalize_name(source)
        visit_source = SOURCE_MAP.get(source_str, 'online_etc') if source_str else ''
        
        area_str = normalize_name(area)
        treatment_area = AREA_MAP.get(area_str, 'etc') if area_str else ''
        
        staff_name = normalize_name(staff)
        referrer_name = normalize_name(referrer)
        
        records.append((
            f"pt-{uuid.uuid4().hex[:8]}",  # id
            HOSPITAL_ID,
            '',  # chart_number
            patient_name,
            '',  # phone
            '',  # birth_date
            '',  # gender
            'new',  # patient_type
            visit_source,
            source_str,  # visit_source_detail
            referrer_name,
            visit_date,  # first_visit_date
            visit_date,  # last_visit_date
            1,  # visit_count
            treatment_area,
            '',  # primary_doctor
            staff_name,  # assigned_counselor
            staff_name,  # desk_staff
            '',  # visit_reason
            '',  # address
            '',  # addr_sido
            '',  # addr_sigungu
            '',  # addr_detail
            '',  # memo
            'active',
            '',  # kakao_registered
            'import-crm'  # created_by
        ))
    
    print(f"✅ 파싱 완료: {len(records)}건 (스킵: {skipped}건)")
    
    # 통계
    source_cnt = Counter(r[8] for r in records if r[8])
    area_cnt = Counter(r[14] for r in records if r[14])
    year_cnt = Counter(r[11][:4] for r in records if r[11])
    staff_cnt = Counter(r[16] for r in records if r[16])
    
    print(f"\n📊 내원경로 분포:")
    for k, v in source_cnt.most_common():
        print(f"  {k}: {v}")
    print(f"\n📊 진료영역 분포:")
    for k, v in area_cnt.most_common():
        print(f"  {k}: {v}")
    print(f"\n📊 연도 분포:")
    for k, v in sorted(year_cnt.items()):
        print(f"  {k}: {v}")
    print(f"\n📊 접수자 Top 10:")
    for k, v in staff_cnt.most_common(10):
        print(f"  {k}: {v}")
    
    # ═══ SQLite 직접 삽입 ═══
    print(f"\n🔄 SQLite 직접 삽입 시작...")
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()
    
    # 기존 import 데이터 삭제
    cur.execute(f"DELETE FROM patients WHERE hospital_id=? AND created_by='import-crm'", (HOSPITAL_ID,))
    deleted = cur.rowcount
    print(f"  기존 데이터 {deleted}건 삭제")
    
    # 배치 INSERT
    sql = """INSERT INTO patients (id, hospital_id, chart_number, patient_name, phone, birth_date, gender,
        patient_type, visit_source, visit_source_detail, referrer_name,
        first_visit_date, last_visit_date, visit_count,
        treatment_area, primary_doctor, assigned_counselor, desk_staff,
        visit_reason, address, addr_sido, addr_sigungu, addr_detail, memo, status, kakao_registered, created_by)
        VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)"""
    
    batch_size = 500
    total = 0
    errors = 0
    
    for i in range(0, len(records), batch_size):
        batch = records[i:i+batch_size]
        try:
            cur.executemany(sql, batch)
            total += len(batch)
        except Exception as e:
            # 개별 삽입 시도
            for r in batch:
                try:
                    cur.execute(sql, r)
                    total += 1
                except Exception as e2:
                    errors += 1
                    if errors <= 5:
                        print(f"  ⚠️ 에러: {r[3]} - {e2}")
        
        if (i // batch_size) % 5 == 0:
            print(f"  진행: {min(i + batch_size, len(records))}/{len(records)}")
    
    conn.commit()
    
    # 검증
    cur.execute(f"SELECT COUNT(*) FROM patients WHERE hospital_id=?", (HOSPITAL_ID,))
    total_in_db = cur.fetchone()[0]
    
    cur.execute(f"SELECT COUNT(*) FROM patients WHERE hospital_id=? AND created_by='import-crm'", (HOSPITAL_ID,))
    imported = cur.fetchone()[0]
    
    cur.execute(f"SELECT visit_source, COUNT(*) as c FROM patients WHERE hospital_id=? AND created_by='import-crm' GROUP BY visit_source ORDER BY c DESC", (HOSPITAL_ID,))
    print(f"\n📋 DB 내원경로별 검증:")
    for row in cur.fetchall():
        print(f"  {row[0] or '(빈값)'}: {row[1]}")
    
    cur.execute(f"SELECT treatment_area, COUNT(*) as c FROM patients WHERE hospital_id=? AND created_by='import-crm' GROUP BY treatment_area ORDER BY c DESC", (HOSPITAL_ID,))
    print(f"\n📋 DB 진료영역별 검증:")
    for row in cur.fetchall():
        print(f"  {row[0] or '(빈값)'}: {row[1]}")
    
    conn.close()
    
    print(f"\n🎉 Import 완료!")
    print(f"  삽입: {total}건, 에러: {errors}건")
    print(f"  DB 총 환자수: {total_in_db}건 (import: {imported}건)")

if __name__ == '__main__':
    main()
