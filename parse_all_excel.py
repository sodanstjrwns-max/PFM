#!/usr/bin/env python3
"""
종합 엑셀 파싱 스크립트
5개 엑셀 파일의 모든 시트를 파싱하여 SQL INSERT 문으로 변환
hospital_id: 34653f75-cb75-4b73-b52c-d5675c83bb9f (sbddc@naver.com)
"""

import openpyxl
import os
import uuid
import json
import re
from datetime import datetime, date

HOSPITAL_ID = '34653f75-cb75-4b73-b52c-d5675c83bb9f'
USER_ID = '34a05d21-c7db-4a95-9b10-7a721b331dec'  # sbddc@naver.com

def esc(val):
    """Escape string for SQL"""
    if val is None:
        return 'NULL'
    s = str(val).strip()
    if not s or s == 'None' or s == 'nan':
        return 'NULL'
    s = s.replace("'", "''")
    return f"'{s}'"

def uid():
    return str(uuid.uuid4())[:8]

def safe_row(row, idx, default=None):
    """Safely access row index"""
    try:
        return row[idx] if idx < len(row) else default
    except (IndexError, TypeError):
        return default


def parse_date(val):
    """Parse various date formats"""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime('%Y-%m-%d')
    if isinstance(val, date):
        return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if not s or s == 'None':
        return None
    # Handle 2025.01.03 format
    s = s.replace('.', '-')
    # Handle YYYY-MM-DD HH:MM:SS
    m = re.match(r'(\d{4})-(\d{1,2})-(\d{1,2})', s)
    if m:
        return f"{m.group(1)}-{int(m.group(2)):02d}-{int(m.group(3)):02d}"
    return None

def parse_number(val):
    """Parse number from cell"""
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if val != val:  # NaN check
            return None
        return val
    s = str(val).strip().replace(',', '').replace('만', '0000')
    try:
        return float(s)
    except:
        return None

def get_excel_files():
    """Get all xlsx files by size to avoid encoding issues"""
    fdir = '/home/user/uploaded_files'
    files = {}
    for fn in os.listdir(fdir):
        if not fn.endswith('.xlsx'):
            continue
        fp = os.path.join(fdir, fn)
        size = os.path.getsize(fp)
        # Identify by known sizes
        if size == 2071674:
            files['CRM'] = fp
        elif size == 390711:
            files['KPI'] = fp
        elif size == 5488786:
            files['NOTE'] = fp
        elif size == 1357546:
            files['BOGO'] = fp
        elif size == 689476:
            files['HOSPITAL'] = fp
    return files

def read_sheet_data(wb, sheet_name, header_row=0, max_rows=None):
    """Read sheet data as list of dicts"""
    if sheet_name not in wb.sheetnames:
        return []
    ws = wb[sheet_name]
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if max_rows and i > max_rows + header_row:
            break
        rows.append(list(row))
    if len(rows) <= header_row:
        return []
    headers = [str(h).strip() if h else f'col_{i}' for i, h in enumerate(rows[header_row])]
    data = []
    for row in rows[header_row+1:]:
        if all(v is None for v in row):
            continue
        d = {}
        for j, h in enumerate(headers):
            d[h] = row[j] if j < len(row) else None
        data.append(d)
    return data

###############################################################################
# 1. CRM-2.xlsx parsers
###############################################################################

def parse_crm_inbound(wb):
    """콜-인바운드 -> call_records (call_type='inbound')"""
    data = read_sheet_data(wb, '콜-인바운드')
    sqls = []
    for row in data:
        call_date = parse_date(row.get('', row.get('col_2')))  # 3rd column is date
        if not call_date:
            # Try from column values  
            year = row.get('연도')
            month = row.get('월')
            # The date column might be unnamed
            for k, v in row.items():
                d = parse_date(v)
                if d and d.startswith('20'):
                    call_date = d
                    break
            if not call_date:
                continue
        
        name = row.get('환자분 성함', '')
        phone = str(row.get('연락처', '')).strip() if row.get('연락처') else ''
        patient_type = str(row.get('신/구환', '')).strip()
        staff = str(row.get('상담원', '')).strip()
        treatment = str(row.get('관심 진료', '')).strip()
        recog = str(row.get('인지경로', '')).strip()
        res_status = str(row.get('예약여부', '')).strip()
        res_date = parse_date(row.get('예약일'))
        res_fulfilled = str(row.get('예약이행여부', '')).strip() if row.get('예약이행여부') else ''
        follow_up = str(row.get('예약미이행f/u', '')).strip() if row.get('예약미이행f/u') else ''
        
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
            
        rid = f"cr-in-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO call_records (id, hospital_id, call_type, call_date, patient_name, phone, patient_type, staff_name, treatment_interest, recognition_path, reservation_status, reservation_date, reservation_fulfilled, follow_up, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, 'inbound', {esc(call_date)}, {esc(name)}, {esc(phone)}, {esc(patient_type)}, {esc(staff)}, {esc(treatment)}, {esc(recog)}, {esc(res_status)}, {esc(res_date)}, {esc(res_fulfilled)}, {esc(follow_up)}, {esc(USER_ID)});"
        )
    return sqls

def parse_crm_outbound(wb):
    """콜-아웃바운드 -> call_records (call_type='outbound')"""
    data = read_sheet_data(wb, '콜-아웃바운드')
    sqls = []
    for row in data:
        call_date = parse_date(row.get('날짜'))
        phone = str(row.get('연락처', '')).strip() if row.get('연락처') else ''
        name = row.get('환자명', '')
        patient_type = str(row.get('신/구환', '')).strip()
        staff = str(row.get('응대자', '')).strip()
        treatment = str(row.get('관심 진료', '')).strip()
        recog = str(row.get('인지경로', '')).strip()
        res_status = str(row.get('예약여부', '')).strip()
        res_date = parse_date(row.get('예약일'))
        comment = str(row.get('코멘트', '')).strip() if row.get('코멘트') else ''
        
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
            
        rid = f"cr-out-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO call_records (id, hospital_id, call_type, call_date, patient_name, phone, patient_type, staff_name, treatment_interest, recognition_path, reservation_status, reservation_date, comment, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, 'outbound', {esc(call_date)}, {esc(name)}, {esc(phone)}, {esc(patient_type)}, {esc(staff)}, {esc(treatment)}, {esc(recog)}, {esc(res_status)}, {esc(res_date)}, {esc(comment)}, {esc(USER_ID)});"
        )
    return sqls

def parse_crm_new_patient_route(wb):
    """신환 내원 경로 개별 -> patients table"""
    data = read_sheet_data(wb, '신환 내원 경로 개별')
    sqls = []
    for row in data:
        visit_date = parse_date(row.get('날짜'))
        name = row.get('성함', '')
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
        staff = str(row.get('접수자', '')).strip()
        treatment = str(row.get('진료 영역', '')).strip()
        source = str(row.get('내원 경로', '')).strip()
        referrer = str(row.get('소개자', '')).strip() if row.get('소개자') else ''
        
        rid = f"pt-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patients (id, hospital_id, patient_name, patient_type, visit_source, visit_source_detail, referrer_name, first_visit_date, last_visit_date, treatment_area, assigned_counselor, status, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(name)}, '신환', {esc(source)}, {esc(source)}, {esc(referrer)}, {esc(visit_date)}, {esc(visit_date)}, {esc(treatment)}, {esc(staff)}, 'active', {esc(USER_ID)});"
        )
    return sqls

def parse_crm_complaints_detail(wb):
    """컴플레인 개별 -> complaints"""
    data = read_sheet_data(wb, '컴플레인 개별')
    sqls = []
    for row in data:
        cdate = parse_date(row.get('날짜'))
        if not cdate:
            continue
        name = row.get('환자 성함', '')
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
        responder = str(row.get('응대자', '')).strip()
        part = str(row.get('파트', '')).strip()
        detail = str(row.get('세부 내용', '')).strip()
        resolver = str(row.get('해결자', '')).strip()
        resolution = str(row.get('해결 내용', '')).strip()
        summary = str(row.get('내용 정리', '')).strip()
        
        rid = f"cmp-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO complaints (id, hospital_id, complaint_date, patient_name, part, category, description, responder, resolver, resolution, status, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(cdate)}, {esc(name)}, {esc(part)}, {esc(part)}, {esc(detail or summary)}, {esc(responder)}, {esc(resolver)}, {esc(resolution)}, 'resolved', {esc(USER_ID)});"
        )
    return sqls

def parse_crm_receivables(wb):
    """미수금 리스트 -> consult_records (as financial tracking)"""
    data = read_sheet_data(wb, '미수금 리스트')
    sqls = []
    for row in data:
        rdate = parse_date(row.get('날짜'))
        if not rdate:
            continue
        name = row.get('성함', '')
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
        chart_no = str(row.get('차트번호', '')).strip()
        staff = str(row.get('응대자', '')).strip()
        amount = parse_number(row.get('미수금'))
        memo = str(row.get('리콜종결 문자 여부  o,x', '')).strip() if row.get('리콜종결 문자 여부  o,x') else ''
        
        rid = f"rcv-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO consult_records (id, hospital_id, record_date, patient_name, counselor_name, agreed_amount, notes, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(name)}, {esc(staff)}, {amount if amount else 'NULL'}, {esc('미수금: ' + memo)}, {esc(USER_ID)});"
        )
    return sqls

def parse_crm_daily_check(wb):
    """데일리 체크 -> reservation_records + wait_time_records + parking_records"""
    ws = wb['데일리 체크 '] if '데일리 체크 ' in wb.sheetnames else None
    if not ws:
        return []
    
    sqls = []
    # Complex layout: row 1 has dates, columns have different metrics
    # Already parsed in previous seed - skip if counts are sufficient
    # We'll read this more carefully
    rows = []
    for row in ws.iter_rows(values_only=True):
        rows.append(list(row))
    
    if len(rows) < 3:
        return sqls
    
    # Find date row and data rows
    for i, row in enumerate(rows):
        if i == 0:
            continue  # header
        if not row or len(row) < 1:
            continue
        date_val = parse_date(row[0]) if row[0] else None
        if not date_val:
            continue
        
        if len(row) < 3:
            continue
        day_of_week = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        cancel = parse_number(row[2]) if len(row) > 2 else None
        dentweb_cancel = parse_number(row[3]) if len(row) > 3 else None
        fulfill_rate = parse_number(row[4]) if len(row) > 4 else None
        total_wait = parse_number(row[5]) if len(row) > 5 else None
        avg_wait = parse_number(row[6]) if len(row) > 6 else None
        parking = parse_number(row[7]) if len(row) > 7 else None
        
        if cancel is not None or dentweb_cancel is not None:
            rid = f"rr-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO reservation_records (id, hospital_id, record_date, day_of_week, cancel_count, dentweb_cancel_count, fulfillment_rate, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(date_val)}, {esc(day_of_week)}, {int(cancel) if cancel else 0}, {int(dentweb_cancel) if dentweb_cancel else 0}, {fulfill_rate if fulfill_rate else 'NULL'}, {esc(USER_ID)});"
            )
        
        if total_wait is not None or avg_wait is not None:
            rid = f"wt-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO wait_time_records (id, hospital_id, record_date, day_of_week, total_wait_minutes, avg_wait_minutes, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(date_val)}, {esc(day_of_week)}, {total_wait if total_wait else 0}, {avg_wait if avg_wait else 0}, {esc(USER_ID)});"
            )
        
        if parking is not None:
            rid = f"pk-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO parking_records (id, hospital_id, record_date, day_of_week, ticket_count, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(date_val)}, {esc(day_of_week)}, {int(parking)}, {esc(USER_ID)});"
            )
    
    return sqls

def parse_crm_insurance_check(wb):
    """보험진료체크 -> daily_records (insurance data supplement)"""
    data = read_sheet_data(wb, '보험진료체크')
    sqls = []
    for row in data:
        rdate = parse_date(row.get('날짜'))
        if not rdate:
            continue
        total_patients = parse_number(row.get('총 환자수'))
        zero_patients = parse_number(row.get('수납 0인 환자수'))
        ratio = parse_number(row.get('비율'))
        
        if total_patients is not None:
            rid = f"dr-ins-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO daily_records (id, hospital_id, record_date, existing_patients, notes, recorded_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {int(total_patients)}, {esc(f'보험체크: 총환자 {int(total_patients)}, 수납0 {int(zero_patients) if zero_patients else 0}, 비율 {ratio:.1%}' if ratio else f'보험체크: 총환자 {int(total_patients)}')}, {esc(USER_ID)});"
            )
    return sqls

def parse_crm_reappointment(wb):
    """실장님 재예약 확인 환자 리스트 -> patient_funnel"""
    data = read_sheet_data(wb, '실장님 재예약 확인 환자 리스트')
    sqls = []
    for row in data:
        name = row.get('환자명', '')
        if not name or str(name).strip() in ('None', '', 'nan'):
            continue
        staff = str(row.get('담당자', '')).strip()
        situation = str(row.get('중단상황', '')).strip()
        recall_count = str(row.get('리콜횟수', '')).strip()
        last_recall = str(row.get('마지막 리콜', '')).strip()
        
        rid = f"pf-reappt-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patient_funnel (id, hospital_id, patient_name, current_stage, treatment_type, assigned_doctor, notes) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(name)}, 'recall', '재예약', {esc(staff)}, {esc(f'중단: {situation}, 리콜횟수: {recall_count}, 마지막리콜: {last_recall}')});"
        )
    return sqls

###############################################################################
# 2. 월간 KPI-2.xlsx parsers  
###############################################################################

def parse_kpi_monthly(wb):
    """월별 시트들 (23년8월 ~ 26년3월) -> kpi_targets + daily_records"""
    sqls = []
    
    for sn in wb.sheetnames:
        # Match patterns like '25년 3월', '24년12월', '23년 8월'
        m = re.match(r'(\d{2})년\s*(\d{1,2})월', sn)
        if not m:
            continue
        
        yr = int(m.group(1)) + 2000
        mo = int(m.group(2))
        year_month = f"{yr}-{mo:02d}"
        
        ws = wb[sn]
        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append(list(row))
        
        if len(rows) < 2:
            continue
        
        # First row header has target info, data rows have daily values
        # Column layout: date, day, 목표매출, 실제매출, 차이, 차이누계, 누적매출, 비급여목표, 비급여실제, ...
        
        monthly_target = None
        monthly_actual = 0
        monthly_insurance_actual = 0
        daily_count = 0
        
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            rdate = parse_date(row[0])
            if not rdate:
                continue
            
            day_of_week = str(row[1]).strip() if row[1] else ''
            target = parse_number(row[2])
            actual = parse_number(row[3])
            insur_target = parse_number(row[7])
            insur_actual = parse_number(row[8]) if len(row) > 8 else None
            
            if actual is not None and actual > 0:
                monthly_actual += actual
                daily_count += 1
                
                # Daily record
                rid = f"dr-kpi-{uid()}"
                non_insur = actual
                insur = insur_actual if insur_actual else 0
                sqls.append(
                    f"INSERT OR IGNORE INTO daily_records (id, hospital_id, record_date, day_of_week, revenue_non_insurance, revenue_insurance, notes, recorded_by) "
                    f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(day_of_week)}, {non_insur}, {insur}, {esc(f'KPI 매출: 목표 {target}, 실제 {actual}')}, {esc(USER_ID)});"
                )
            
            if insur_actual is not None:
                monthly_insurance_actual += insur_actual
        
        # Monthly KPI target
        if daily_count > 0:
            # Extract monthly totals from header area
            rid = f"kpi-{uid()}"
            # Get header row info for target  
            header = rows[0]
            total_target = parse_number(header[2]) if len(header) > 2 else None
            
            sqls.append(
                f"INSERT OR IGNORE INTO kpi_targets (id, hospital_id, year_month, target_revenue, weekdays, notes, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(year_month)}, {monthly_actual}, {daily_count}, {esc(f'{sn} - 실제매출합계: {monthly_actual}')}, {esc(USER_ID)});"
            )
    
    return sqls

def parse_kpi_monthly_comparison(wb):
    """월별 페이스 비교 -> kpi summary data for kpi_targets"""
    data = read_sheet_data(wb, '월별 페이스 비교')
    sqls = []
    # This is a comparison chart - headers are dates, rows are values
    # Better captured by the monthly sheets above
    return sqls

###############################################################################
# 3. 실장 노트-2.xlsx parsers (5.4MB - large file, read-only mode)
###############################################################################

def parse_note_consult_records(wb):
    """월별 상담 기록 시트들 (263, 262, 261, 2512, ...) -> consult_records"""
    sqls = []
    consult_sheets = []
    for sn in wb.sheetnames:
        # Match patterns: 263, 262, 2512, 2511, 2510, ... (YYMM format)
        if re.match(r'^2[3-6]\d{1,2}$', sn):
            consult_sheets.append(sn)
        elif sn in ['6월 임시']:
            consult_sheets.append(sn)
    
    for sn in consult_sheets:
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 200:  # Limit per sheet
                break
            rows.append(list(row))
        
        if len(rows) < 2:
            continue
        
        # Headers: 날짜, 챠트번호, 성함, 상담의, 상담사, (총상담금액/비용계획), 동의금액, (수납금액), 할인내역, (구/신환), 진료카테고리, 치료확정 ...
        headers = [str(h).strip() if h else '' for h in rows[0]]
        
        for row in rows[1:]:
            if all(v is None for v in row[:5]):
                continue
            
            rdate = parse_date(row[0])
            chart_no = str(row[1]).strip() if row[1] else ''
            name = str(row[2]).strip() if row[2] else ''
            
            if not name or name in ('None', '', 'nan'):
                continue
            
            doctor = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            counselor = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            
            # Find amount columns - they vary by sheet
            agreed_amount = None
            planned_amount = None
            patient_type = ''
            category = ''
            confirmed = ''
            discount = ''
            
            for j, h in enumerate(headers):
                if j >= len(row):
                    break
                val = row[j]
                if '동의금액' in h:
                    agreed_amount = parse_number(val)
                elif '총상담금액' in h or '비용계획' in h:
                    planned_amount = parse_number(val)
                elif '구/신환' in h or '구신환' in h:
                    patient_type = str(val).strip() if val else ''
                elif '진료 카테고리' in h or '진료카테고리' in h:
                    category = str(val).strip() if val else ''
                elif '치료확정' in h:
                    confirmed = str(val).strip() if val else ''
                elif '할인' in h:
                    discount = str(val).strip() if val else ''
            
            rid = f"cslt-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO consult_records (id, hospital_id, record_date, chart_number, patient_name, doctor_name, counselor_name, planned_amount, agreed_amount, discount_note, patient_type, treatment_category, treatment_confirmed, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(chart_no)}, {esc(name)}, {esc(doctor)}, {esc(counselor)}, {planned_amount if planned_amount else 'NULL'}, {agreed_amount if agreed_amount else 'NULL'}, {esc(discount)}, {esc(patient_type)}, {esc(category)}, {esc(confirmed)}, {esc(USER_ID)});"
            )
    
    return sqls

def parse_note_proactive_calls(wb):
    """선제적 통화 -> call_records (proactive type)"""
    ws = wb['선제적 통화'] if '선제적 통화' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    for row in rows[1:]:
        if all(v is None for v in row[:4]):
            continue
        call_date = parse_date(row[0])
        staff = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        summary = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        reaction = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        followup = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"cr-pro-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO call_records (id, hospital_id, call_type, call_date, patient_name, staff_name, call_purpose, comment, follow_up, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, 'outbound', {esc(call_date)}, {esc(name)}, {esc(staff)}, '선제적 통화', {esc(summary + ' / 반응: ' + reaction)}, {esc(followup)}, {esc(USER_ID)});"
        )
    return sqls

def parse_note_outbound_calls(wb):
    """아웃바운드콜 -> call_records"""
    ws = wb['아웃바운드콜'] if '아웃바운드콜' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # 날짜, 성함, 전화번호, 자연유입/광고유입, 전화유무, 예약유무, 예약날짜, 상담사, 상담금액, 진료금액
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        call_date = parse_date(row[0])
        name = str(row[1]).strip() if row[1] else ''
        phone = str(row[2]).strip() if row[2] else ''
        source = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        called = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        reserved = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        res_date = parse_date(row[6]) if len(row) > 6 else None
        counselor = str(row[7]).strip() if len(row) > 7 and row[7] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"cr-ob-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO call_records (id, hospital_id, call_type, call_date, patient_name, phone, staff_name, recognition_path, reservation_status, reservation_date, comment, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, 'outbound', {esc(call_date)}, {esc(name)}, {esc(phone)}, {esc(counselor)}, {esc(source)}, {esc(reserved)}, {esc(res_date)}, {esc(f'전화: {called}')}, {esc(USER_ID)});"
        )
    return sqls

def parse_note_referrals(wb):
    """소개 -> patients (referral tracking)"""
    ws = wb['소개'] if '소개' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # 날짜, 소개해주신 분, 담당자, 소개받고 오신분, 감사 문자 여부, 감사 전화 여부, 환자 반응
    for row in rows[1:]:
        if all(v is None for v in row[:4]):
            continue
        rdate = parse_date(row[0])
        referrer = str(row[1]).strip() if row[1] else ''
        staff = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        new_patient = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        
        if not new_patient or new_patient in ('None', '', 'nan'):
            continue
        
        rid = f"pt-ref-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patients (id, hospital_id, patient_name, patient_type, visit_source, referrer_name, first_visit_date, assigned_counselor, status, memo, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(new_patient)}, '신환', '소개', {esc(referrer)}, {esc(rdate)}, {esc(staff)}, 'active', '소개환자', {esc(USER_ID)});"
        )
    return sqls

def parse_note_unconfirmed(wb):
    """미확정환자관리 -> patient_funnel"""
    ws = None
    for sn in wb.sheetnames:
        if '미확정환자' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    for row in rows[1:]:
        if all(v is None for v in row[:4]):
            continue
        consult_date = parse_date(row[0])
        chart_no = str(row[1]).strip() if row[1] else ''
        name = str(row[2]).strip() if row[2] else ''
        counselor = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        confirmed = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        feedback = str(row[5]).strip() if len(row) > 5 and row[5] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"pf-unc-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patient_funnel (id, hospital_id, patient_name, current_stage, assigned_doctor, notes) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(name)}, 'unconfirmed', {esc(counselor)}, {esc(f'확정: {confirmed}, 피드백: {feedback}')});"
        )
    return sqls

def parse_note_ortho_stats(wb):
    """교정월별결과 -> consult_records (orthodontic summary)"""
    ws = None
    for sn in wb.sheetnames:
        if '교정월별결과' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 100:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # 해당월, 전체상담, 미진단, 진단, 진행O, 진행X, 성장f/u, 매출, 진단후동의율
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        month = str(row[0]).strip() if row[0] else ''
        total = parse_number(row[1])
        undiag = parse_number(row[2])
        diag = parse_number(row[3]) if len(row) > 3 else None
        proceed_yes = parse_number(row[4]) if len(row) > 4 else None
        proceed_no = parse_number(row[5]) if len(row) > 5 else None
        revenue = parse_number(row[7]) if len(row) > 7 else None
        agree_rate = parse_number(row[8]) if len(row) > 8 else None
        
        if not month or total is None:
            continue
        
        notes = f"교정 월별결과 {month}: 전체상담 {int(total) if total else 0}, 진단 {int(diag) if diag else 0}, 진행 {int(proceed_yes) if proceed_yes else 0}, 매출 {revenue if revenue else 0}"
        rid = f"cslt-ort-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO consult_records (id, hospital_id, record_date, patient_name, treatment_category, notes, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(parse_date(month) or '2025-01-01')}, '교정 월간 요약', '교정', {esc(notes)}, {esc(USER_ID)});"
        )
    return sqls

def parse_note_counselor_records(wb):
    """개별 상담사 시트 (강혜란, 홍서영, 황혜인, 박설희, 조아영, 한혜림) -> consult_records"""
    counselor_sheets = ['강혜란', '홍서영', '황혜인', '박설희', '조아영', '한혜림']
    sqls = []
    
    for sn in counselor_sheets:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 200:
                break
            rows.append(list(row))
        
        if len(rows) < 2:
            continue
        
        headers = [str(h).strip() if h else '' for h in rows[0]]
        
        for row in rows[1:]:
            if all(v is None for v in row[:4]):
                continue
            
            rdate = parse_date(row[0])
            name = ''
            chart_no = ''
            category = ''
            confirmed = ''
            
            for j, h in enumerate(headers):
                if j >= len(row):
                    break
                val = row[j]
                if val is None:
                    continue
                if '성함' in h:
                    name = str(val).strip()
                elif '챠트번호' in h or '차트번호' in h:
                    chart_no = str(val).strip()
                elif '진료 카테고리' in h:
                    category = str(val).strip()
                elif '치료확정' in h or '확정' in h:
                    confirmed = str(val).strip()
            
            if not name or name in ('None', '', 'nan'):
                continue
            
            rid = f"cslt-{sn[:2]}-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO consult_records (id, hospital_id, record_date, chart_number, patient_name, counselor_name, treatment_category, treatment_confirmed, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(chart_no)}, {esc(name)}, {esc(sn)}, {esc(category)}, {esc(confirmed)}, {esc(USER_ID)});"
            )
    
    return sqls

###############################################################################
# 4. 보고리스트.xlsx parsers
###############################################################################

def parse_bogo_team_report(wb):
    """팀팀장보고리스트 -> meetings (as team reports)"""
    data = read_sheet_data(wb, '팀팀장보고리스트')
    sqls = []
    for row in data[:200]:  # Limit
        rdate = parse_date(row.get(list(row.keys())[0] if row else ''))
        reporter = str(list(row.values())[1]).strip() if len(row) > 1 and list(row.values())[1] else ''
        if not reporter or reporter in ('None', '', 'nan'):
            continue
        
        # Gather report content from remaining columns
        content_parts = []
        for k, v in list(row.items())[2:]:
            if v and str(v).strip() not in ('None', '', 'nan'):
                content_parts.append(str(v).strip())
        content = ' | '.join(content_parts[:5])
        
        if not content:
            continue
        
        rid = f"mtg-team-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO meetings (id, hospital_id, title, description, meeting_date, status, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(f'팀장보고 - {reporter}')}, {esc(content)}, {esc(rdate or '2026-03-01')}, 'completed', {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_morning_report(wb):
    """아침보고리스트 -> meetings"""
    ws = wb['아침보고리스트'] if '아침보고리스트' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        rdate = parse_date(row[1]) if len(row) > 1 else None
        if not rdate:
            continue
        content_parts = [str(v).strip() for v in row[2:] if v and str(v).strip() not in ('None', '', 'nan')]
        if not content_parts:
            continue
        
        rid = f"mtg-am-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO meetings (id, hospital_id, title, description, meeting_date, status, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, '아침보고', {esc(' | '.join(content_parts[:5]))}, {esc(rdate)}, 'completed', {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_staff_referral(wb):
    """직원소개 환자 -> patients"""
    data = read_sheet_data(wb, '직원소개 환자')
    sqls = []
    for row in data:
        rdate = parse_date(row.get('날짜이', row.get(list(row.keys())[0])))
        staff = str(row.get('소개 직원', '')).strip()
        name = str(row.get('소개 환자 성함', '')).strip()
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"pt-stref-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patients (id, hospital_id, patient_name, patient_type, visit_source, referrer_name, first_visit_date, status, memo, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(name)}, '신환', '직원소개', {esc(staff)}, {esc(rdate)}, 'active', '직원소개환자', {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_complaint_patients(wb):
    """컴플레인 환자 -> complaints"""
    ws = None
    for sn in wb.sheetnames:
        if '컴플레인 환자' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        rdate = parse_date(row[1]) if len(row) > 1 else None
        name = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        responder = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        content = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"cmp-bg-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO complaints (id, hospital_id, complaint_date, patient_name, description, responder, status, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(name)}, {esc(content)}, {esc(responder)}, 'resolved', {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_review_event(wb):
    """데스크 리뷰 이벤트 -> review_management"""
    ws = wb['데스크 리뷰 이벤트'] if '데스크 리뷰 이벤트' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:  # Limit - this sheet has 5971 rows
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        reviewer_staff = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"rv-desk-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO review_management (id, hospital_id, platform, reviewer_name, rating, review_text, review_date, sentiment, response_status, registered_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, '네이버', {esc(name)}, 5, {esc(f'리뷰 권유: {reviewer_staff}')}, datetime('now'), 'positive', 'pending', {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_new_email(wb):
    """신환이메일 -> patients (email data)"""
    ws = None
    for sn in wb.sheetnames:
        if '신환이메일' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        rdate = parse_date(row[0])
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        email = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"pt-em-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO patients (id, hospital_id, patient_name, patient_type, first_visit_date, status, memo, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(name)}, '신환', {esc(rdate)}, 'active', {esc(f'이메일: {email}')}, {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_finish_patients(wb):
    """마무리환자리스트 -> call_records (follow-up calls)"""
    ws = wb['마무리환자리스트'] if '마무리환자리스트' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:  # Limit - 5356 rows
            break
        rows.append(list(row))
    
    for row in rows[1:]:
        if all(v is None for v in row[:3]):
            continue
        rdate = parse_date(row[0])
        name = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        staff = str(row[2]).strip() if len(row) > 2 and row[2] else ''
        called = str(row[3]).strip() if len(row) > 3 and row[3] else ''
        satisfied = str(row[4]).strip() if len(row) > 4 and row[4] else ''
        
        if not name or name in ('None', '', 'nan'):
            continue
        
        rid = f"cr-fin-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO call_records (id, hospital_id, call_type, call_date, patient_name, staff_name, call_purpose, comment, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, 'outbound', {esc(rdate)}, {esc(name)}, {esc(staff)}, '마무리전화', {esc(f'전화: {called}, 만족: {satisfied}')}, {esc(USER_ID)});"
        )
    return sqls

def parse_bogo_facilities(wb):
    """시설장비 -> materials (as facility management records)"""
    ws = None
    for sn in wb.sheetnames:
        if '시설장비' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    if len(rows) < 3:
        return sqls
    
    # Complex header - skip for now, just record basics
    for row in rows[2:]:
        if all(v is None for v in row[:3]):
            continue
        item = str(row[0]).strip() if row[0] else ''
        if not item or item in ('None', '', 'nan', '시설/장비 관리 대장'):
            continue
        
        rid = f"mat-fac-{uid()}"
        details = ' | '.join([str(v).strip() for v in row[1:8] if v and str(v).strip() not in ('None', '', 'nan')])
        sqls.append(
            f"INSERT OR IGNORE INTO materials (id, hospital_id, title, description, file_type) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(item)}, {esc(details)}, 'facility');"
        )
    return sqls

def parse_bogo_leave_management(wb):
    """전직원 연차관리표 + 휴가관리표 -> leave_balances"""
    sqls = []
    for sn in wb.sheetnames:
        if '연차관리표' in sn or '휴가관리표' in sn:
            ws = wb[sn]
            rows = []
            for i, row in enumerate(ws.iter_rows(values_only=True)):
                if i > 100:
                    break
                rows.append(list(row))
            
            if len(rows) < 3:
                continue
            
            # These are complex formatted sheets - extract key data
            for row in rows[1:]:
                if all(v is None for v in row[:3]):
                    continue
                name = str(row[0]).strip() if row[0] else ''
                if not name or name in ('None', '', 'nan', '실장.', '실장'):
                    continue
                
                # Try to extract leave data
                total = parse_number(row[1]) if len(row) > 1 else None
                used = parse_number(row[2]) if len(row) > 2 else None
                
                if total is not None:
                    rid = f"lb-{uid()}"
                    sqls.append(
                        f"INSERT OR IGNORE INTO leave_balances (id, hospital_id, user_id, year, leave_type, total_days, used_days) "
                        f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(USER_ID)}, 2026, 'annual', {total}, {used if used else 0});"
                    )
    return sqls

###############################################################################
# 5. 병원관리.xlsx parsers
###############################################################################

def parse_hospital_weekly(wb):
    """주간 -> daily_records (comprehensive weekly dashboard data)"""
    ws = wb['주간'] if '주간' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # Headers: 병원관리 통합 대시보드, 날짜, 매출, 보험청구액, 총액, 보험비중, 일평균매출, 매출증감율, 신환수, ...
    headers = [str(h).strip() if h else '' for h in rows[0]]
    
    for row in rows[1:]:
        if not row or len(row) < 3:
            continue
        rdate = parse_date(row[1]) if len(row) > 1 else None
        if not rdate:
            continue
        
        revenue = parse_number(row[2]) if len(row) > 2 else None
        insurance = parse_number(row[3]) if len(row) > 3 else None
        new_patients = parse_number(row[8]) if len(row) > 8 else None
        
        if revenue is None and new_patients is None:
            continue
        
        rid = f"dr-wk-{uid()}"
        sqls.append(
            f"INSERT OR IGNORE INTO daily_records (id, hospital_id, record_date, revenue_non_insurance, revenue_insurance, new_patients, recorded_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {revenue if revenue else 0}, {insurance if insurance else 0}, {int(new_patients) if new_patients else 0}, {esc(USER_ID)});"
        )
    return sqls

def parse_hospital_monthly_stats(wb):
    """월별 통계 -> kpi_targets"""
    ws = wb['월별 통계'] if '월별 통계' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # Headers: '', 신환, 구환, 총환, 일평균신환, 일평균내원, 매출, 공단청구, 하루매출
    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        period = str(row[0]).strip() if row[0] else ''
        if not period or period in ('None', '', 'nan'):
            continue
        
        new_patients = parse_number(row[1]) if len(row) > 1 else None
        existing = parse_number(row[2]) if len(row) > 2 else None
        total = parse_number(row[3]) if len(row) > 3 else None
        avg_new = parse_number(row[4]) if len(row) > 4 else None
        avg_visit = parse_number(row[5]) if len(row) > 5 else None
        revenue = parse_number(row[6]) if len(row) > 6 else None
        insurance = parse_number(row[7]) if len(row) > 7 else None
        
        if revenue is None and new_patients is None:
            continue
        
        notes = f"월별통계 {period}: 신환 {int(new_patients) if new_patients else 0}, 구환 {int(existing) if existing else 0}, 매출 {revenue if revenue else 0}"
        rid = f"kpi-ms-{uid()}"
        
        # Try to parse year-month from period
        ym = parse_date(period)
        if ym:
            ym = ym[:7]  # YYYY-MM
        else:
            ym = period
        
        sqls.append(
            f"INSERT OR IGNORE INTO kpi_targets (id, hospital_id, year_month, target_revenue, target_new_patients_weekday, notes, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(ym)}, {revenue if revenue else 0}, {int(new_patients) if new_patients else 0}, {esc(notes)}, {esc(USER_ID)});"
        )
    return sqls

def parse_hospital_new_patient_mgmt(wb):
    """신환 관리, 주별 신환 관리 -> patients (aggregated)"""
    sqls = []
    for sn in ['신환 관리', '주별 신환 관리']:
        if sn not in wb.sheetnames:
            continue
        ws = wb[sn]
        rows = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i > 200:
                break
            rows.append(list(row))
        
        if len(rows) < 2:
            continue
        
        # Headers: '', '', 총신환수, 임플란트신환수, 임플비율, 교정신환수, 교정비율, 소아신환수, 소아비율, ...
        for row in rows[1:]:
            if not row or len(row) < 2:
                continue
            period = str(row[0]).strip() if row[0] else ''
            if not period or period in ('None', '', 'nan'):
                continue
            total = parse_number(row[2]) if len(row) > 2 else None
            implant = parse_number(row[3]) if len(row) > 3 else None
            ortho = parse_number(row[5]) if len(row) > 5 else None
            pedo = parse_number(row[7]) if len(row) > 7 else None
            
            if total is None:
                continue
            
            notes = f"신환관리 {period}: 총 {int(total)}, 임플 {int(implant) if implant else 0}, 교정 {int(ortho) if ortho else 0}, 소아 {int(pedo) if pedo else 0}"
            rid = f"kpi-np-{uid()}"
            rdate = parse_date(period)
            if rdate:
                ym = rdate[:7]
            else:
                ym = period
            
            sqls.append(
                f"INSERT OR IGNORE INTO kpi_targets (id, hospital_id, year_month, target_new_patients_weekday, notes, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(ym)}, {int(total)}, {esc(notes)}, {esc(USER_ID)});"
            )
    return sqls

def parse_hospital_reservation_cancel(wb):
    """예약취소 월별 -> reservation_records"""
    ws = None
    for sn in wb.sheetnames:
        if '예약취소' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 500:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # 날짜, 요일, 예약취소수, 덴트웹취소수, 예약이행율, 총대기시간, 평균대기시간, 주차권총수
    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        rdate = parse_date(row[0])
        if not rdate:
            continue
        day = str(row[1]).strip() if len(row) > 1 and row[1] else ''
        cancel = parse_number(row[2]) if len(row) > 2 else None
        dw_cancel = parse_number(row[3]) if len(row) > 3 else None
        rate = parse_number(row[4]) if len(row) > 4 else None
        total_wait = parse_number(row[5]) if len(row) > 5 else None
        avg_wait = parse_number(row[6]) if len(row) > 6 else None
        parking = parse_number(row[7]) if len(row) > 7 else None
        
        if cancel is not None:
            rid = f"rr-hm-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO reservation_records (id, hospital_id, record_date, day_of_week, cancel_count, dentweb_cancel_count, fulfillment_rate, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(day)}, {int(cancel)}, {int(dw_cancel) if dw_cancel else 0}, {rate if rate else 'NULL'}, {esc(USER_ID)});"
            )
        if total_wait is not None or avg_wait is not None:
            rid = f"wt-hm-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO wait_time_records (id, hospital_id, record_date, day_of_week, total_wait_minutes, avg_wait_minutes, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(day)}, {total_wait if total_wait else 0}, {avg_wait if avg_wait else 0}, {esc(USER_ID)});"
            )
        if parking is not None:
            rid = f"pk-hm-{uid()}"
            sqls.append(
                f"INSERT OR IGNORE INTO parking_records (id, hospital_id, record_date, day_of_week, ticket_count, created_by) "
                f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(day)}, {int(parking)}, {esc(USER_ID)});"
            )
    
    return sqls

def parse_hospital_implant_stats(wb):
    """임플란트 통계 -> daily_records (implant supplement)"""
    ws = None
    for sn in wb.sheetnames:
        if '임플란트 통계' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # '', 식립수, 어벗수, 오스템, 포인트, 스트라우만, 보험, 오스템비율, 보험비율
    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        period = str(row[0]).strip() if row[0] else ''
        if not period or period in ('None', '', 'nan'):
            continue
        install = parse_number(row[1]) if len(row) > 1 else None
        abutment = parse_number(row[2]) if len(row) > 2 else None
        
        if install is None:
            continue
        
        notes = f"임플란트 통계 {period}: 식립 {int(install)}, 어벗 {int(abutment) if abutment else 0}"
        rid = f"dr-impl-{uid()}"
        rdate = parse_date(period) or '2025-01-01'
        
        sqls.append(
            f"INSERT OR IGNORE INTO daily_records (id, hospital_id, record_date, notes, recorded_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate)}, {esc(notes)}, {esc(USER_ID)});"
        )
    return sqls

def parse_hospital_naver_reviews(wb):
    """네이버리뷰&유튜브 -> review_management"""
    ws = None
    for sn in wb.sheetnames:
        if '네이버리뷰' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    # This is likely a comparison/tracking sheet, not individual reviews
    # Headers: ㅡ, 더스퀘어, 더서울, 라이크, 비디, ...
    # Skip detailed parsing, this is aggregate data
    return sqls

def parse_hospital_marketing(wb):
    """마케팅 -> marketing_records"""
    ws = wb['마케팅'] if '마케팅' in wb.sheetnames else None
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 200:
            break
        rows.append(list(row))
    
    if len(rows) < 2:
        return sqls
    
    # 기간, 신환, 홈페이지, 체류시간, 스마트플레이스, 예약유입, 예약신청, 이용완료, 예약리뷰
    for row in rows[1:]:
        if not row or len(row) < 2:
            continue
        period = str(row[0]).strip() if row[0] else ''
        if not period or period in ('None', '', 'nan'):
            continue
        new_patients = parse_number(row[1]) if len(row) > 1 else None
        
        if new_patients is None:
            continue
        
        # Create marketing record
        rid = f"mk-{uid()}"
        rdate = parse_date(period) or '2025-01-01'
        
        notes_parts = []
        labels = ['신환', '홈페이지', '체류시간', '스마트플레이스', '예약유입', '예약신청', '이용완료', '예약리뷰']
        for j, label in enumerate(labels):
            if j + 1 < len(row) and row[j + 1] is not None:
                notes_parts.append(f"{label}: {row[j+1]}")
        
        sqls.append(
            f"INSERT OR IGNORE INTO marketing_records (id, hospital_id, record_month, new_patients, created_at) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate[:7] if rdate else period)}, {int(new_patients)}, datetime('now'));"
        )
    return sqls

def parse_hospital_hr(wb):
    """HR -> attendance / staff tracking"""
    ws = wb['HR'] if 'HR' in wb.sheetnames else None
    if not ws:
        return []
    # HR data is complex formatted - skip for initial import
    return []

def parse_hospital_visit_route(wb):
    """내원 경로 진짜 -> patients (visit source summary)"""
    ws = None
    for sn in wb.sheetnames:
        if '내원 경로 진짜' in sn:
            ws = wb[sn]
            break
    if not ws:
        return []
    sqls = []
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i > 100:
            break
        rows.append(list(row))
    
    # This is summary data - better to record as kpi_targets notes
    if len(rows) < 2:
        return sqls
    
    # '', 날짜, 총신환수, 지인소개, 환자소개, 직원소개, 전체소개, 소개비율, 홈페이지, ...
    for row in rows[1:]:
        if not row or len(row) < 3:
            continue
        rdate = parse_date(row[1]) if len(row) > 1 else None
        total = parse_number(row[2]) if len(row) > 2 else None
        
        if not rdate or total is None:
            continue
        
        referral = parse_number(row[6]) if len(row) > 6 else 0
        referral_rate = parse_number(row[7]) if len(row) > 7 else 0
        
        rid = f"kpi-vr-{uid()}"
        notes = f"내원경로: 총신환 {int(total)}, 소개 {int(referral) if referral else 0}, 소개율 {referral_rate:.1%}" if referral_rate else f"내원경로: 총신환 {int(total)}"
        sqls.append(
            f"INSERT OR IGNORE INTO kpi_targets (id, hospital_id, year_month, target_new_patients_weekday, notes, created_by) "
            f"VALUES ({esc(rid)}, {esc(HOSPITAL_ID)}, {esc(rdate[:7])}, {int(total)}, {esc(notes)}, {esc(USER_ID)});"
        )
    return sqls


###############################################################################
# Main execution
###############################################################################

def main():
    files = get_excel_files()
    print(f"Found files: {list(files.keys())}")
    
    all_sqls = []
    all_sqls.append("-- Full Excel Data Import for sbddc@naver.com")
    all_sqls.append(f"-- Hospital ID: {HOSPITAL_ID}")
    all_sqls.append(f"-- Generated: {datetime.now().isoformat()}")
    all_sqls.append("")
    
    # First, delete existing imported data to avoid duplicates
    all_sqls.append("-- Clear previously imported Excel data")
    tables_to_clear = [
        'call_records', 'complaints', 'daily_records', 'reservation_records',
        'wait_time_records', 'parking_records', 'consultations', 'consult_records',
        'patients', 'patient_funnel', 'kpi_targets', 'meetings', 'marketing_records',
        'review_management', 'materials', 'leave_balances'
    ]
    for t in tables_to_clear:
        all_sqls.append(f"DELETE FROM {t} WHERE hospital_id = '{HOSPITAL_ID}';")
    all_sqls.append("")
    
    # Parse CRM-2.xlsx
    if 'CRM' in files:
        print("\n=== Parsing CRM-2.xlsx ===")
        wb = openpyxl.load_workbook(files['CRM'], read_only=True, data_only=True)
        
        try:

        
            sqls = parse_crm_inbound(wb)

        
            print(f"  콜-인바운드: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_crm_outbound(wb)

        
            print(f"  콜-아웃바운드: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_crm_new_patient_route(wb)

        
            print(f"  신환 내원 경로: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_crm_complaints_detail(wb)

        
            print(f"  컴플레인 개별: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_crm_receivables(wb)

        
            print(f"  미수금 리스트: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:
            try:

                sqls = parse_crm_daily_check(wb)

                print(f"  데일리 체크: {len(sqls)} records")

                all_sqls.extend(sqls)

            except Exception as e:

                print(f"  ERROR: {e}")
        except Exception as e:
            print(f"  데일리 체크: ERROR - {e}")
        
        try:

        
            sqls = parse_crm_insurance_check(wb)

        
            print(f"  보험진료체크: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_crm_reappointment(wb)

        
            print(f"  재예약 확인: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        wb.close()
    
    # Parse KPI-2.xlsx
    if 'KPI' in files:
        print("\n=== Parsing KPI-2.xlsx ===")
        wb = openpyxl.load_workbook(files['KPI'], read_only=True, data_only=True)
        
        try:

        
            sqls = parse_kpi_monthly(wb)

        
            print(f"  월별 KPI: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        wb.close()
    
    # Parse 실장 노트-2.xlsx
    if 'NOTE' in files:
        print("\n=== Parsing 실장 노트-2.xlsx ===")
        wb = openpyxl.load_workbook(files['NOTE'], read_only=True, data_only=True)
        
        try:

        
            sqls = parse_note_consult_records(wb)

        
            print(f"  상담 기록: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_proactive_calls(wb)

        
            print(f"  선제적 통화: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_outbound_calls(wb)

        
            print(f"  아웃바운드콜: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_referrals(wb)

        
            print(f"  소개: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_unconfirmed(wb)

        
            print(f"  미확정환자: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_ortho_stats(wb)

        
            print(f"  교정월별결과: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_note_counselor_records(wb)

        
            print(f"  상담사별 기록: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        wb.close()
    
    # Parse 보고리스트.xlsx
    if 'BOGO' in files:
        print("\n=== Parsing 보고리스트.xlsx ===")
        wb = openpyxl.load_workbook(files['BOGO'], data_only=True)
        
        try:

        
            sqls = parse_bogo_team_report(wb)

        
            print(f"  팀장보고: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_morning_report(wb)

        
            print(f"  아침보고: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_staff_referral(wb)

        
            print(f"  직원소개환자: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_complaint_patients(wb)

        
            print(f"  컴플레인 환자: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_review_event(wb)

        
            print(f"  리뷰 이벤트: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_new_email(wb)

        
            print(f"  신환이메일: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_finish_patients(wb)

        
            print(f"  마무리환자: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_facilities(wb)

        
            print(f"  시설장비: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_bogo_leave_management(wb)

        
            print(f"  연차/휴가: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        wb.close()
    
    # Parse 병원관리.xlsx
    if 'HOSPITAL' in files:
        print("\n=== Parsing 병원관리.xlsx ===")
        wb = openpyxl.load_workbook(files['HOSPITAL'], data_only=True)
        
        try:

        
            sqls = parse_hospital_weekly(wb)

        
            print(f"  주간 대시보드: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_monthly_stats(wb)

        
            print(f"  월별 통계: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_new_patient_mgmt(wb)

        
            print(f"  신환 관리: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_reservation_cancel(wb)

        
            print(f"  예약취소: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_implant_stats(wb)

        
            print(f"  임플란트 통계: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_marketing(wb)

        
            print(f"  마케팅: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        try:

        
            sqls = parse_hospital_visit_route(wb)

        
            print(f"  내원 경로: {len(sqls)} records")

        
            all_sqls.extend(sqls)

        
        except Exception as e:

        
            print(f"  ERROR: {e}")
        
        wb.close()
    
    # Write output
    output_path = '/home/user/webapp/seed-excel-full.sql'
    with open(output_path, 'w') as f:
        f.write('\n'.join(all_sqls))
    
    total_inserts = sum(1 for s in all_sqls if s.startswith('INSERT'))
    total_deletes = sum(1 for s in all_sqls if s.startswith('DELETE'))
    print(f"\n{'='*60}")
    print(f"✅ Output: {output_path}")
    print(f"   Total INSERT: {total_inserts}")
    print(f"   Total DELETE: {total_deletes}")
    print(f"   File size: {os.path.getsize(output_path)//1024}KB")

if __name__ == '__main__':
    main()
